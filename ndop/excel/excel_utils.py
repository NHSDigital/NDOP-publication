import pandas as pd
from ndop.config import config, params
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from typing import Tuple, Any
from pathlib import Path


def get_excel_template():
    return os.path.join(params.ROOT_DIR, params.TEMPLATES_FOLDER, params.EXCEL_TEMPLATE)

def get_nhs_logo():
    return Path(params.ROOT_DIR) / params.INPUTS_FOLDER / params.NHS_LOGO_IMG

def write_table_to_sheet(
    wb: openpyxl.Workbook, table_data: pd.DataFrame, sheet_name: str
) -> openpyxl.Workbook:
    """Given some data, a workbook, and a sheet name; writes that data to the chosen sheet

    Args:
        wb (openpyxl.Workbook): The workbook
        table_data (pd.DataFrame): The data to write
        sheet_name (str): The sheet to write to

    Returns:
        openpyxl.Workbook: The workbook, with the data written
    """
    start_cell = find_cell_by_tag(wb, sheet_name, "<start>")
    end_cell = find_cell_by_tag(wb, sheet_name, "<end>")

    ws = wb[sheet_name]
    write_df_from_start_cell(
        start_cell=start_cell, end_cell=end_cell, ws=ws, df=table_data
    )
    return wb


def write_df_from_start_cell(
    start_cell: Tuple, end_cell: Tuple, ws: openpyxl.worksheet, df: pd.DataFrame
) -> None:
    """Given a pandas dataframe and a worksheet, writes that dataframe to that worksheet. Starts at the cell with the <start> tag. Any empty rows after the last df row are written, and up to the <end> cell row, are deleted.

    Args:
        start_cell (Tuple): Cell to start the data in
        end_cell (Tuple): Cell to delete blank rows up to
        ws (openpyxl.worksheet): The worksheet to write to
        df (pd.DataFrame): The data to write
    """
    rows_to_write = dataframe_to_rows(df, index=False, header=False)
    loc = list(start_cell)
    for row in rows_to_write:
        for cell in row:
            ws.cell(row=loc[0], column=loc[1]).value = cell
            loc[1] += 1
        loc[0] += 1
        loc[1] = start_cell[1]
    clear_empty_rows(ws=ws, last_written_row=loc[0], end_cell=end_cell)


def clear_empty_rows(
    ws: openpyxl.worksheet, last_written_row: int, end_cell: Tuple
) -> None:
    """Deletes blank / empty rows from the area where data is written to. This is a workaround for a problem with working from templates; how do we size the template without knowing how many rows the df will have? To get round this, we insert many many blank rows, and then delete the unused ones.

    Args:
        ws (openpyxl.worksheet): the worksheet to clear empty rows from
        last_written_row (int): The last row to which we wrote data
        end_cell (Tuple): The location of the <end> tag
    """
    number_to_delete = end_cell[0] - last_written_row
    ws.delete_rows(last_written_row + 1, number_to_delete)


def find_cell_by_tag(
    wb: openpyxl.Workbook, sheet: openpyxl.worksheet, tag: str
) -> Tuple:
    """Given a tag and a sheet name, findes the index of that the cell which has that tag

    Args:
        wb (openpyxl.Workbook):
        sheet (openpyxl.worksheet):
        tag (str):

    Returns:
        index of cell containing tag
    """
    max_search = 1000
    ws = wb[sheet]
    iter = 0
    while iter <= max_search:
        for row in ws.iter_rows():
            for cell in row:
                iter += 1
                if cell.value == tag:
                    return openpyxl.utils.cell.coordinate_to_tuple(cell.coordinate)
    return None


def write_single_val(ws: openpyxl.worksheet, column: str, tag: str, val: Any, formatting = None, merge = None, alignment = False, row_height = None) -> None:
    """
    Writes a single value into the passed worksheet, identifying the cell using the passed column and tag.

    Args:
        ws(openpyxl.Worksheet): The worksheet to edit
        column(str): The specific column to search
        tag(str): The tag by which to locate the specific cell in the column
        val(float): The value to write to the cell
    """
    loc = find_cell_in_column(ws=ws, tag=tag, column=column)
    ws.cell(row=loc[0], column=loc[1]).value = val

    if formatting: 
        ws.cell(row=loc[0], column=loc[1]).font = formatting
    
    if merge: 
        ws.merge_cells(start_row = loc[0], end_row = loc[0], start_column = loc[1], end_column = merge)

    if alignment == True: 
        ws.cell(row=loc[0], column=loc[1]).alignment = Alignment(wrapText=True)
    
    if row_height:
        ws.row_dimensions[loc[0]].height = row_height

    return None


def find_cell_in_column(ws: openpyxl.worksheet, tag: str, column: str) -> Tuple:
    """Finds a cell with a tag in a particular column

    Args:
        ws (openpyxl.worksheet): The worksheet
        tag (str): The tag in the ell
        column (str): The column you're looking in, A or B etc.

    Returns:
        Tuple: The index of the cell containing the tag
    """
    max_search = 1000
    iter = 0
    while iter <= max_search:
        for cell in ws[column]:
            iter += 1
            if cell.value == tag:
                return openpyxl.utils.cell.coordinate_to_tuple(cell.coordinate)
    return None


def add_nhs_logo_to_sheet(ws: openpyxl.Workbook.worksheets) -> None: 
    """Adds NHS logo to top corner of sheet.

    Args:
        ws (openpyxl.Workbook.worksheets): worksheet to add logo to.

    Returns:
        None.
    """
    ws_max = get_column_letter(ws.max_column)
    img = openpyxl.drawing.image.Image(get_nhs_logo())
    img.anchor = f'{ws_max}1'

    ws.add_image(img)

    return None

